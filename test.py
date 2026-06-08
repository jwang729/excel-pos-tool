import os
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import load_workbook


# =========================================================
# 設定區
# =========================================================

# 從第幾列開始處理
START_ROW = 2

# 黃色代碼
YELLOW_COLORS = {
    "FFFF00",
    "FFFFFF00",
}


# =========================================================
# 判斷 Cell 是否為黃色填滿
# =========================================================
def is_yellow_fill(cell):

    fill = cell.fill

    if fill is None:
        return False

    try:
        color = fill.fgColor.rgb

        if color is None:
            return False

        color = str(color).upper()

        if color in YELLOW_COLORS:
            return True

    except:
        pass

    return False


# =========================================================
# 找最右側第一個空白欄
# =========================================================
def get_first_empty_column(ws):

    max_col = ws.max_column

    # 從右往左找最後有值的欄位
    for col in range(max_col, 0, -1):

        has_value = False

        for row in range(1, ws.max_row + 1):

            value = ws.cell(row=row, column=col).value

            if value not in [None, ""]:
                has_value = True
                break

        if has_value:
            return col + 1

    return max_col + 1


# =========================================================
# 處理 Excel
# =========================================================
def process_excel(file_path):

    wb = load_workbook(file_path)

    # 每個 Sheet 處理
    for ws in wb.worksheets:

        max_row = ws.max_row
        max_col = ws.max_column

        # 找最右側第一個空白欄
        target_col = get_first_empty_column(ws)

        # 欄位名稱固定為 FLAG
        ws.cell(row=1, column=target_col).value = "FLAG"

        # 每一列
        for row in range(START_ROW, max_row + 1):

            yellow_count = 0

            # 掃描所有 Cell
            for col in range(1, max_col + 1):

                cell = ws.cell(row=row, column=col)

                if is_yellow_fill(cell):
                    yellow_count += 1

            target_cell = ws.cell(row=row, column=target_col)

            # 判斷規則
            if yellow_count >= 3:
                target_cell.value = "A"

            elif yellow_count >= 1:
                target_cell.value = "M"

        print(f"Sheet 完成: {ws.title}")

    # =====================================================
    # 產出檔名
    # =====================================================
    original_path = Path(file_path)

    output_file = (
        original_path.parent /
        f"{original_path.stem}_finished.xlsx"
    )

    wb.save(output_file)

    return str(output_file)


# =========================================================
# GUI 選檔
# =========================================================
def choose_file():

    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(
        title="選擇 Excel 檔案",
        filetypes=[
            ("Excel Files", "*.xlsx *.xlsm"),
            ("All Files", "*.*")
        ]
    )

    if not file_path:
        return

    try:

        output_path = process_excel(file_path)

        messagebox.showinfo(
            "完成",
            f"處理完成！\n\n輸出檔案：\n{output_path}"
        )

    except Exception as e:

        messagebox.showerror(
            "錯誤",
            str(e)
        )


# =========================================================
# 主程式
# =========================================================
if __name__ == "__main__":

    choose_file()