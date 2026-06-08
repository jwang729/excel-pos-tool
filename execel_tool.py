import openpyxl
import os
import time
import sys

# =============================
# 判斷 cell 是否有填色
# =============================
def is_cell_filled(cell):
    fill = cell.fill
    if fill is None:
        return False

    if fill.fill_type is not None:
        if fill.start_color.rgb and fill.start_color.rgb != "00000000":
            return True

    return False


# =============================
# 處理單一 Excel
# =============================
def process_excel(file_path):

    print(f"\n📄 處理檔案: {file_path}")

    start_time = time.time()
    wb = openpyxl.load_workbook(file_path)

    total_rows = sum(ws.max_row - 1 for ws in wb.worksheets if ws.max_row > 1)
    processed_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        print(f"➡️ 工作表: {sheet_name}")

        for row in ws.iter_rows(min_row=2):
            filled_count = sum(1 for cell in row if is_cell_filled(cell))

            if filled_count > 0:
                for cell in row:
                    if cell.value is None or str(cell.value).strip() == "":
                        cell.value = "A" if filled_count >= 3 else "M"
                        break

            processed_rows += 1

            # ✅ CLI進度顯示
            if processed_rows % 100 == 0:
                print(f"   ⏳ 已處理 {processed_rows}/{total_rows} rows")

    # ✅ 輸出
    base, _ = os.path.splitext(file_path)
    new_file = base + "_processed.xlsx"
    wb.save(new_file)

    elapsed = round(time.time() - start_time, 2)

    print(f"✅ 完成: {new_file} (耗時 {elapsed} 秒)")


# =============================
# 主程式入口
# =============================
def main():

    if len(sys.argv) < 2:
        print("❗ 使用方式:")
        print("python excel_tool.py <檔案路徑 或 資料夾路徑>")
        return

    path = sys.argv[1]

    # ✅ 單一檔案
    if os.path.isfile(path):
        process_excel(path)

    # ✅ 資料夾
    elif os.path.isdir(path):
        files = [
            os.path.join(path, f)
            for f in os.listdir(path)
            if f.endswith(".xlsx")
        ]

        print(f"\n📂 共發現 {len(files)} 個 Excel 檔案")

        for file in files:
            process_excel(file)

    else:
        print("❌ 路徑不存在")


if __name__ == "__main__":
    main()
    
    